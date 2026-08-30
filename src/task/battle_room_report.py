# -*- coding: utf-8 -*-
"""
战备房间测试执行报告生成器

同一份结果输出两种形态：
  - JSON  report_<ts>.json           机器可读（含 UI 快照、待支持清单）
  - XLSX  report_<ts>.xlsx           人读/QA 归档，列结构与
            《搜打撤测试用例整合.xlsx》一致，便于回填
"""

import datetime
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.task.battle_room_checks import NEEDS_SUPPORT

_STATUS_FILL = {
    'PASS': PatternFill('solid', fgColor='C6EFCE'),
    'FAIL': PatternFill('solid', fgColor='FFC7CE'),
    'Blocked': PatternFill('solid', fgColor='FFEB9C'),
    'N/A': PatternFill('solid', fgColor='DDDDDD'),
}
_HEADER_FILL = PatternFill('solid', fgColor='4472C4')
_THIN = Side(style='thin', color='B0B0B0')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def load_env_info():
    """
    读取测试环境信息（全局设置"测试环境"）

    Returns:
        dict: {'目标环境': str, '私服服务器目录': str}，读取失败给默认值
    """
    import config as app_config
    env = {'目标环境': '未配置', '私服服务器目录': ''}
    try:
        path = app_config.get_config_path('测试环境.json')
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, dict):
            for k in list(env):
                if data.get(k):
                    env[k] = data[k]
    except (OSError, ValueError):
        pass
    return env


def write_reports(out_dir, version, smoke, aborted, results, summary,
                  ui_snapshot, needs_support_extra=None, environment=None,
                  dismissed_popups=None):
    """
    写 JSON + XLSX 两份报告

    Returns:
        (json_path, xlsx_path)
    """
    os.makedirs(out_dir, exist_ok=True)
    ts = f'{datetime.datetime.now():%Y%m%d_%H%M%S}'
    json_path = os.path.join(out_dir, f'report_{ts}.json')
    xlsx_path = os.path.join(out_dir, f'report_{ts}.xlsx')

    needs_support = needs_support_extra
    if needs_support is None:
        needs_support = [
            {'case_id': r.case.case_id, 'priority': r.case.priority,
             'feature': r.case.feature,
             'reason': NEEDS_SUPPORT.get(r.case.case_id, r.detail)}
            for r in results if r.status == 'N/A'
        ]
    if environment is None:
        environment = {}

    report = {
        'generated_at': datetime.datetime.now().isoformat(),
        'version': version,
        'environment': environment,
        'smoke': smoke,
        'aborted': aborted,
        'summary': summary,
        'cases': [r.to_dict() for r in results],
        'needs_unity_support': needs_support,
        'dismissed_popups': dismissed_popups or [],
        'ui_snapshot': ui_snapshot[:300],
    }
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    _write_xlsx(xlsx_path, version, smoke, aborted, results, summary,
                needs_support, ui_snapshot, environment, dismissed_popups)
    return json_path, xlsx_path


def _style_header(ws, headers):
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
    ws.freeze_panes = 'A2'


def _autosize(ws, max_width=60):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value) if cell.value else ''
            widths[cell.column] = min(max(widths.get(cell.column, 10),
                                          len(v) + 2), max_width)
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w


def _write_xlsx(path, version, smoke, aborted, results, summary,
                needs_support, ui_snapshot, environment=None,
                dismissed_popups=None):
    wb = Workbook()
    wb.properties.creator = 'ok-jump BattleRoomTestTask'

    # ---- Sheet1 汇总 ----
    ws = wb.active
    ws.title = '汇总'
    executed = summary['PASS'] + summary['FAIL'] + summary['Blocked']
    rate = (summary['PASS'] / executed * 100) if executed else 0.0
    rows = [
        ('报告时间', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('工具版本', version),
        ('目标环境', (environment or {}).get('目标环境', '')),
        ('私服服务器目录', (environment or {}).get('私服服务器目录', '')),
        ('冒烟模式(仅P0)', '是' if smoke else '否'),
        ('用例总数', len(results)),
        ('PASS', summary['PASS']),
        ('FAIL', summary['FAIL']),
        ('Blocked', summary['Blocked']),
        ('N/A(暂不可自动化)', summary['N/A']),
        ('Skipped', summary['Skipped']),
        ('可执行通过率', f'{rate:.0f}%'),
        ('提前终止', '是' if aborted else '否'),
        ('UI快照节点数', len(ui_snapshot)),
        ('自动跳过的弹窗', ' ; '.join(dismissed_popups or []) or '无'),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    _autosize(ws)

    # ---- Sheet2 用例结果（列结构对齐 QA xlsx）----
    ws2 = wb.create_sheet('用例结果')
    headers = ['用例编号', '优先级', '模块', '功能', '操作步骤', '期望结果',
               '测试结果', '耗时(秒)', '执行详情']
    _style_header(ws2, headers)
    wrap = Alignment(wrap_text=True, vertical='top')
    for i, r in enumerate(results, 2):
        c = r.case
        vals = [c.case_id, c.priority, c.module, c.feature, c.steps,
                c.expected, r.status, round(r.duration, 2), r.detail]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.alignment = wrap
            cell.border = _BORDER
            if col == 7:
                fill = _STATUS_FILL.get(r.status)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)
    _autosize(ws2)

    # ---- Sheet3 待 Unity 支持 ----
    ws3 = wb.create_sheet('待Unity侧支持')
    _style_header(ws3, ['用例编号', '优先级', '功能', '自动化缺口'])
    for i, item in enumerate(needs_support, 2):
        for col, key in enumerate(('case_id', 'priority', 'feature', 'reason'), 1):
            cell = ws3.cell(row=i, column=col, value=item.get(key))
            cell.alignment = wrap
            cell.border = _BORDER
    _autosize(ws3)

    wb.save(path)
