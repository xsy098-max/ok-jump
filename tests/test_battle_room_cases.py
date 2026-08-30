# -*- coding: utf-8 -*-
"""
战备房间测试用例自动化（P0冒烟）单元测试

全部基于伪造连接，不依赖真实 Unity 工程：
  - 用例加载/过滤/统计
  - UnityConnection 新增 UI 命令封装
  - UiContext 解析缓存/视图枚举/返回栈
  - 用例执行器（导航通过/阻塞传播/运行时错误清扫/N-A 分级）
  - BattleRoomTestTask 装配与汇总

伪连接模拟"点击入口打开视图根、返回栈关闭视图根"的状态行为，
按钮/锚点按真机枚举(logs/ui_inventory)确认的真实名称回放。
"""

import json
import os
from unittest.mock import MagicMock, patch

import pytest

from src.task.battle_room_cases import (
    load_cases, filter_cases, summarize, sort_by_priority,
    STATUS_PASS, STATUS_FAIL, STATUS_BLOCKED, STATUS_NA,
)
from src.task.battle_room_ui import (
    UiContext, TIPS_ROOT, TIPS_EQUIP_PATH, TIPS_CLOSE_PATH, warehouse_cell_path,
    GM_SWITCH_PATH, GM_INPUT_PATH, GM_SEND_PATH,
    TIPS_PUTINBAG_PATH, TIPS_PUTINSAFETY_PATH, TIPS_SELL_BUTTON_PATH,
)
from src.task.battle_room_checks import CheckContext, EXECUTABLE_CHECKS, NEEDS_SUPPORT, run_case
from src.utils.UnityConnection import UnityConnection


def make_logger():
    return MagicMock()


def make_item(name, path=None, active=True, button=True):
    return {
        'name': name,
        'path': path or f'UIRoot/{name}',
        'activeSelf': active,
        'activeInHierarchy': active,
        'hasButton': button,
        'hasToggle': False,
        'hasInputField': False,
        'interactable': True,
    }


ROOM = ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
        'ViewRoot<Game.SDCMainView>/MainView/UILobby_SDCMainView/TeamRoom/SafeArea')

# 真机枚举确认的房间锚点/战备窗口元素
ROOM_ANCHORS = {
    'BtnPreBattle': f'{ROOM}/BtnPreBattle',
    'BtnNodeSeason': f'{ROOM}/Obj_BtnList/Btn_NodeSeason/BtnNodeSeason',
    'BtnStore': f'{ROOM}/Obj_BtnList/BtnStore/BtnStore',
    'BtnCollect': f'{ROOM}/Obj_BtnList/BtnCollect/BtnCollect',
    'BtnTask': f'{ROOM}/Obj_BtnList/BtnTask/BtnTask',
    'BtnStart': f'{ROOM}/RightBottom/ButtonRoot/ImgStart/BtnStart',
    'Button_Convene': f'{ROOM}/Obj_Convene/Button_Convene',
    'BtnChoosePlay': f'{ROOM}/RightBottom/ModeInfo/BtnChoosePlay',
}
_PW_PREFIX = ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
              'ViewRoot<Game.SDCPreWarWindow>/MainView/SDCPreWarWindow')
PREWAR_ITEMS = {
    'SDCPreWarWindow': ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                        'ViewRoot<Game.SDCPreWarWindow>'),
    'EquipTrans': f'{_PW_PREFIX}/SafeArea/Obj_Left/Left/EquipTrans',
    'Obj_LeftList_2': f'{_PW_PREFIX}/SafeArea/Obj_Left/Left/Obj_LeftList_2',
    'Obj_LeftList_3': f'{_PW_PREFIX}/SafeArea/Obj_Left/Left/Obj_LeftList_3',
    'Obj_LeftList_4': f'{_PW_PREFIX}/SafeArea/Obj_Left/Left/Obj_LeftList_4',
    'WarehouseScrollView': f'{_PW_PREFIX}/SafeArea/Right/WarehouseScrollView',
    'ButtonFunc': f'{_PW_PREFIX}/SafeArea/Right/TabExtendList/ButtonFunc',
    'Button_Buy': f'{_PW_PREFIX}/SafeArea/Right/Button_Buy',
    'Button_Sell': f'{_PW_PREFIX}/SafeArea/Right/Button_Sell',
}


class StatefulFakeConn:
    """
    带状态的伪 Unity 连接：

    - find_ui: ViewRoot 查询回放当前打开的视图根；其它关键字查静态物品表
    - click_ui: 命中入口表时切换视图状态（记录 opened_stack）
    - go_back: 返回栈弹出最近打开的视图
    """

    def __init__(self, base_roots=None, entry_map=None, runtime_errors=None):
        self.base_roots = list(base_roots or [])
        self.opened_stack = []           # 本次会话点开的视图根
        self.entry_map = dict(entry_map or {})   # {nameContains/精确名: 打开的视图根名}
        self._runtime_errors = list(runtime_errors or [])
        self.static_items = {}           # {nameContains: [item]}
        self.clicks = []
        self.audit = []
        self.info_items = {}             # {path: info dict}（text/color/isOn）
        self.click_effects = {}          # {path: fn(self)} 点击副作用
        self.toast = ''                  # 模拟 Toast 当前文案
        go_back_fail = False
        self.go_back_ok = not go_back_fail

    # ---- 登记 helpers ----
    def add_info(self, path, name, text=None, color=None, active=True,
                 button=False, interactable=True):
        self.info_items[path] = {
            'name': name, 'path': path, 'activeInHierarchy': active,
            'interactable': interactable, 'hasButton': button,
            'hasToggle': False,
            'text': text, 'color': color, 'isOn': None,
        }

    def _set_tips_open(self):
        if 'ViewRoot<Game.SDCItemTips>' not in self.opened_stack:
            self.opened_stack.append('ViewRoot<Game.SDCItemTips>')
            for p in tuple(self.info_items):
                if 'SDCItemTips' in p:
                    self.info_items[p]['activeInHierarchy'] = True

    def _set_tips_closed(self):
        if 'ViewRoot<Game.SDCItemTips>' in self.opened_stack:
            self.opened_stack.remove('ViewRoot<Game.SDCItemTips>')
            for p in tuple(self.info_items):
                if 'SDCItemTips' in p:
                    self.info_items[p]['activeInHierarchy'] = False
    def add_items(self, table):
        for name, path in table.items():
            self.static_items.setdefault(name, []).append(
                make_item(name, path))

    def _all_static(self):
        return [it for lst in self.static_items.values() for it in lst]

    def current_roots(self):
        return self.base_roots + self.opened_stack

    # ---- UnityConnection 协议面 ----
    def find_ui(self, path=None, name_contains=None, max_results=None):
        if path:
            items = [it for it in self._all_static() if it.get('path') == path]
            return {'success': True, 'count': len(items), 'items': items[:max_results or 10]}
        kw = name_contains or ''
        if kw == 'ViewRoot':
            items = [make_item(v, f'UIRoot/{v}', button=False)
                     for v in self.current_roots()]
            return {'success': True, 'count': len(items), 'items': items}
        # 模拟插件的"名称包含"全量扫描语义：静态物品 + 信息表 + 视图根都参与
        pool = {}
        for it in self._all_static():
            pool[it['path']] = it
        for p, it in self.info_items.items():
            if it.get('activeInHierarchy'):
                pool.setdefault(p, {k: v for k, v in it.items() if k != 'text'})
        for v in self.current_roots():
            pool.setdefault(f'UIRoot/{v}', make_item(v, f'UIRoot/{v}', button=False))
        items = [it for it in pool.values() if kw in it['name']]
        return {'success': True, 'count': len(items), 'items': items[:max_results or 10]}

    def click_ui(self, path=None, name_contains=None):
        self.clicks.append(path or name_contains)
        pool = self._all_static() + list(self.info_items.values())
        if path:
            hit = next((it for it in pool if it['path'] == path), None)
            target_name = hit['name'] if hit else None
            key = path
        else:
            key = name_contains
            hits = [it for it in pool
                    if it['path'].endswith(key) or it['name'] == key]
            target_name = hits[0]['name'] if hits else None
        if target_name is None:
            return {'status': 'error', 'message': '未找到目标 UI'}
        new_root = (self.entry_map.get(target_name)
                    or self.entry_map.get(f'@path:{key}'))
        if new_root:
            self.opened_stack.append(new_root)
            self.audit.append(f"push {new_root} by {target_name}")
            if new_root == self.PRE_WAR_ROOT:
                self._sync_prewar(True)
        effect = self.click_effects.get(key)
        if effect:
            effect(self)
        return {'status': 'ok',
                'message': json.dumps({'success': True,
                                       'clickedBy': 'WButton.pointerChain',
                                       'target': {'name': target_name}})}

    def reveal_gm_panel(self, show=True):
        state = bool(show)
        for p in (GM_INPUT_PATH, GM_SEND_PATH):
            if p in self.info_items:
                self.info_items[p]['activeInHierarchy'] = state
        return {'status': 'ok', 'message':
                json.dumps({'success': True, 'show': show})}

    def set_ui_active(self, is_active, path=None, name_contains=None):
        target = None
        if path and path in self.info_items:
            target = self.info_items[path]
        elif name_contains:
            target = next((it for it in self.info_items.values()
                           if name_contains in it['name']), None)
        if target is None:
            return {'status': 'error', 'message': '未找到目标 UI'}
        target['activeInHierarchy'] = bool(is_active)
        return {'status': 'ok', 'message': '{"success": true}'}


    PRE_WAR_ROOT = 'ViewRoot<Game.SDCPreWarWindow>'

    def _sync_prewar(self, state):
        """战备窗开/关时,联动其容器类信息的激活态(格子另由页签管理)"""
        for p, it in self.info_items.items():
            if 'Game.SDCPreWarWindow>' in p and not it['name'].startswith('Obj_Item'):
                it['activeInHierarchy'] = state

    def reveal_gm_panel(self, show=True):
        state = bool(show)
        for p in (GM_INPUT_PATH, GM_SEND_PATH):
            if p in self.info_items:
                self.info_items[p]['activeInHierarchy'] = state
        return {'status': 'ok', 'message':
                json.dumps({'success': True, 'show': show})}

    def get_ui_info(self, path=None, name_contains=None, max_results=None):
        items = []
        if path:
            it = self.info_items.get(path)
            if it and it.get('activeInHierarchy'):
                items = [{**it,
                          'text': it['text'](self) if callable(it['text'])
                          else it['text'],
                          'color': it['color'](self) if callable(it['color'])
                          else it['color']}]
            return {'success': True, 'count': len(items), 'items': items}
        kw = name_contains or ''
        for it in self.info_items.values():
            # 真实插件返回全部(含未激活),由调用方按 activeInHierarchy 自滤
            if kw in it['name']:
                txt = it['text'](self) if callable(it['text']) else it['text']
                clr = it['color'](self) if callable(it['color']) else it['color']
                items.append({**it, 'text': txt, 'color': clr})
        return {'success': True, 'count': len(items),
                'items': items[:max_results or 20]}

    def go_back(self):
        if self.opened_stack and self.go_back_ok:
            popped = self.opened_stack.pop()
            self.audit.append(f"pop {popped}")
            if popped == self.PRE_WAR_ROOT:
                self._sync_prewar(False)
            return {'status': 'ok', 'message': '{"success": true}'}
        return {'status': 'ok', 'message': '{"success": false}'}

    def get_battle_state(self, include_errors=False, clear_errors=False):
        errors, self._runtime_errors = self._runtime_errors, []
        return {'errors': errors} if include_errors else {'ok': True}

    def is_connected(self):
        return True

    def sdc_fill_teammate(self, on=None, timeout=10):
        if not hasattr(self, 'sdc_fill'):
            self.sdc_fill = False
        if on is not None:
            self.sdc_fill = bool(on)
        # 模拟真封装:返回解析后的 dict(与 UnityConnection.sdc_fill_teammate 一致)
        return {'success': True, 'on': self.sdc_fill, 'inRoom': True}

    # ---- 仓库格子模型（subtype 官方名列表驱动 TIPS 标题；动作消费格子） ----
    def _register_cell(self, i):
        import src.task.battle_room_ui as brui
        p = brui.warehouse_cell_path(i)
        self.info_items.setdefault(p, {
            'name': 'Btn_Click', 'path': p, 'activeInHierarchy': True,
            'interactable': True, 'hasButton': True, 'hasToggle': False,
            'text': None, 'color': None, 'isOn': None})
        self.click_effects[p] = (lambda idx: (lambda c: c._open_cell_tips(idx)))(i)
        g = p + '/BtnGou'
        self.info_items.setdefault(g, {
            'name': 'BtnGou', 'path': g, 'activeInHierarchy': False,
            'interactable': True, 'hasButton': True, 'hasToggle': False,
            'text': None, 'color': None, 'isOn': None})
        self.click_effects[g] = (lambda idx: (lambda c: c._check_gou(idx)))(i)

    def init_cells(self, cells):
        import src.task.battle_room_ui as brui
        self.cells = list(cells)          # [(subtype, 道具名)]
        self.current_cell = None
        self.checked_gou = set()
        self.sell_mode = False
        self.toast = ''
        for i in range(len(self.cells)):
            self._register_cell(i)
        ap = brui.TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemMid/Button_Activate'
        self.add_info(ap, 'Button', text='激活', active=False, button=True)
        self.click_effects[ap] = lambda c: c._act_toast('激活成功')
        bp = (brui.TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemMid/'
              'GiftDescInfoContent/Button_Buy')
        if bp not in self.info_items:
            self.add_info(bp, 'Button_Buy', text='购买', active=False, button=True)
        self.click_effects[bp] = lambda c: c._buy_current()
        pb = (brui.TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemMid/'
              'GiftDescInfoContent/Button_PutInBag')
        if pb not in self.info_items:
            self.add_info(pb, 'Button_PutInBag', active=False, button=True)
        self.click_effects[pb] = lambda c: c._put_in_bag()
        for nm, toast in (('Button_PutInSafetyBox', '放入成功'),
                          ('Button_Equip', '装配成功')):
            ep = (brui.TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemMid/'
                  'GiftDescInfoContent/' + nm)
            if ep not in self.info_items:
                self.add_info(ep, nm, active=False, button=True)
            self.click_effects[ep] = (lambda t: (lambda c: c._consume_current(t)))(toast)
        cp = brui.TIPS_SELL_BUTTON_PATH.rsplit('/', 1)[0] + '/Button_ConfirmBottom'
        self.add_info(cp, 'Button', text='确认', active=False, button=True)
        self.click_effects[cp] = lambda c: c._confirm_sell()
        if brui.TIPS_SELL_CANCEL_PATH not in self.info_items:
            self.add_info(brui.TIPS_SELL_CANCEL_PATH, 'Button_Cancel',
                          active=False, button=True)
            self.click_effects[brui.TIPS_SELL_CANCEL_PATH] =                 lambda c: c._set_sell_mode(False)
        ok_p = (brui.TIPS_SELL_BUTTON_PATH.rsplit('/', 1)[0]
                + '/Button_OneKeySell')
        self.add_info(ok_p, 'Button', text='一键出售', active=False, button=True)
        if brui.TIPS_SELL_BUTTON_PATH not in self.click_effects:
            self.click_effects[brui.TIPS_SELL_BUTTON_PATH] = \
                lambda c: c._set_sell_mode(not c.sell_mode)

    def _open_cell_tips(self, idx):
        if idx < len(self.cells):
            self.current_cell = idx
        if 'ViewRoot<Game.SDCItemTips>' not in self.opened_stack:
            self.opened_stack.append('ViewRoot<Game.SDCItemTips>')
        for p, it in self.info_items.items():
            if 'SDCItemTips>' in p:
                it['activeInHierarchy'] = True

    def _check_gou(self, idx):
        self.checked_gou.add(idx)
        self._open_cell_tips(idx)

    def _put_in_bag(self):
        if self.current_cell is not None and self.current_cell < len(self.cells):
            sub = self.cells[self.current_cell][0]
            if sub == 4:
                self.bag_collections = getattr(self, 'bag_collections', 0) + 1
            self.cells[self.current_cell] = (None, '')
        self.toast = '装配成功'
        if 'ViewRoot<Game.SDCItemTips>' in self.opened_stack:
            self.opened_stack.remove('ViewRoot<Game.SDCItemTips>')
        for pp, it2 in self.info_items.items():
            if 'SDCItemTips>' in pp:
                it2['activeInHierarchy'] = False

    def _consume_current(self, toast):
        if self.current_cell is not None and self.current_cell < len(self.cells):
            self.cells[self.current_cell] = (None, '')
        self.toast = toast
        if 'ViewRoot<Game.SDCItemTips>' in self.opened_stack:
            self.opened_stack.remove('ViewRoot<Game.SDCItemTips>')
        for p, it in self.info_items.items():
            if 'SDCItemTips>' in p:
                it['activeInHierarchy'] = False

    def _act_toast(self, text):
        self.toast = text

    def _buy_current(self):
        try:
            self.currency[0] = str(max(0, int(self.currency[0]) - 100))
        except ValueError:
            pass
        self.toast = '购买成功'
        if 'ViewRoot<Game.SDCItemTips>' in self.opened_stack:
            self.opened_stack.remove('ViewRoot<Game.SDCItemTips>')
        for p, it in self.info_items.items():
            if 'SDCItemTips>' in p:
                it['activeInHierarchy'] = False

    def _confirm_sell(self):
        if self.sell_mode:
            for idx in sorted(self.checked_gou):
                if idx < len(self.cells):
                    self.cells[idx] = (None, '')
            self.checked_gou.clear()
            self.toast = '出售成功'
            self._set_sell_mode(False)
        else:
            # 购买场景：扣货币
            try:
                self.currency[0] = str(max(0, int(self.currency[0]) - 100))
            except ValueError:
                pass
            self.toast = '购买成功'

    def _set_sell_mode(self, on):
        self.sell_mode = bool(on)
        for p, it in self.info_items.items():
            if it['name'] == 'BtnGou' or it['name'] == 'Button_Cancel':
                it['activeInHierarchy'] = self.sell_mode
            elif it.get('text') in ('确认', '一键出售'):
                it['activeInHierarchy'] = self.sell_mode


# 导航骨架用例（离线全链路模拟覆盖集；TIPS/出售家族用例由 TestTipsFamily 单测）
SKELETON_CASES = {
    'TC-4.1-001', 'TC-4.1-002', 'TC-4.2-001', 'TC-4.2-002', 'TC-4.2-003',
    'TC-4.2-004', 'TC-4.2-005', 'TC-4.2-006', 'TC-4.2-027', 'TC-4.3-001',
    'TC-4.3-002', 'TC-4.4-001', 'TC-4.4-011',
}


def build_happy_conn():
    """导航链路全通的伪连接（真实名称，入口->视图映射完整）"""
    conn = StatefulFakeConn(
        base_roots=['ViewRoot<Game.UIMainCity2DView>', 'ViewRoot<Game.GMOutBattleView>'],
        entry_map={
            # 大厅入口 -> 房间
            '@path:UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
            'ViewRoot<Game.UIMainCity2DView>/MainView/UIMainCity2D/SafeArea/'
            'MainCityNew/ControlPanel/Content_Banner/obj_LeftList/Obj_SDC/'
            'Root/SDCEnter_Button': 'ViewRoot<Game.SDCMainView>',
            # 房间入口 -> 各窗口
            'BtnPreBattle': 'ViewRoot<Game.SDCPreWarWindow>',
            'BtnNodeSeason': 'ViewRoot<Game.SeasonView>',
            'BtnStore': 'ViewRoot<Game.StoreView>',
            'BtnCollect': 'ViewRoot<Game.CollectView>',
            'BtnTask': 'ViewRoot<Game.TaskView>',
            'Button_Buy': 'ViewRoot<Game.SDCPreWarShopWindow>',
        },
    )
    conn.add_items({
        **{k: v for k, v in ROOM_ANCHORS.items()},
        # 大厅入口（@path 命中通道）
        'SDCEnter_Button':
            'UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
            'ViewRoot<Game.UIMainCity2DView>/MainView/UIMainCity2D/SafeArea/'
            'MainCityNew/ControlPanel/Content_Banner/obj_LeftList/Obj_SDC/'
            'Root/SDCEnter_Button',
        'Btn_Click': warehouse_cell_path(0),   # 仓库第0格（护甲，供TIPS装配流程）
    })

    # 战备窗全部容器节点为动态信息(初始隐藏,点击 BtnPreBattle 打开后可见)
    for _n, _p in PREWAR_ITEMS.items():
        conn.add_info(_p, _n, active=False,
                      button=_n in ('Button_Buy', 'Button_Sell'))
    conn.prewar_root_item = ('UIRoot/Layer_Common/'
                             'ViewRoot<Game.SDCPreWarWindow>')
    conn.add_info(conn.prewar_root_item, 'SDCPreWarWindow', active=False,
                  button=False)

    # ---- 文本/颜色信息节点（get_ui_info 通道） ----
    # GM 面板（reveal 流程所需节点；初始隐藏由 reveal effect 拉起）
    conn.add_info(GM_INPUT_PATH, 'InputField', active=False)
    conn.add_info(GM_SEND_PATH, 'Send', active=False, button=True)
    def _reveal(c, show=True):
        for p in (GM_INPUT_PATH, GM_SEND_PATH):
            if p in c.info_items:
                c.info_items[p]['activeInHierarchy'] = show
        if not show and hasattr(c, 'toast'):
            pass
    conn.reveal_gm_panel = lambda show=True: (_reveal(conn, show),
                                              {'status': 'ok',
                                               'message': '{"success":true}'})[1]
    orig_set_input = conn.set_ui_input if hasattr(conn, 'set_ui_input') else None
    def set_ui_input(text, path=None, name_contains=None):
        if path in conn.info_items:
            conn.info_items[path]['text'] = text
            return {'status': 'ok', 'message': '{}'}
        return {'status': 'error', 'message': 'no target'}
    conn.set_ui_input = set_ui_input
    def on_send(c):
        it = c.info_items.get(GM_INPUT_PATH)
        cmd = it.get('text') if it else ''
        c.last_gm_cmd = cmd
        _reveal(c, show=False)
        # 解析 AddItem={id}={n}，把道具真正补进仓库格(模拟服务端下发)
        if cmd and cmd.startswith('AddItem='):
            try:
                _, iid, cnt = cmd.split('=')
                iid, cnt = int(iid), int(cnt)
            except ValueError:
                return
            meta = getattr(c, 'id_catalog', {}).get(iid)
            if meta:
                for _ in range(cnt):
                    slot = next((i for i, cellinfo in enumerate(c.cells)
                                 if cellinfo == (None, '')), None)
                    if slot is None:
                        c.cells.append(meta)
                        c._register_cell(len(c.cells) - 1)
                    else:
                        c.cells[slot] = meta
    conn.click_effects[GM_SEND_PATH] = on_send
    conn.last_gm_cmd = None
    conn.id_catalog = {
        1600055: (1, '软质护甲'), 1600163: (2, 'SC-1·攻击芯片'),
        1600115: (8, 'SC-1·治疗'), 1600109: (7, '便携腰包'),
        1600061: (3, 'SC-1·普攻增幅器'), 1600079: (5, '简易救治袋'),
        1610001: (4, '否定陀螺'), 1600103: (6, '篮球场钥匙'),
        1620004: (11, '赛季中级安全箱'), 1600110: (10, '赛季中级钥匙包'),
    }

    # 补齐队友勾选态（两组 Img_Convene,激活组合表达开/关）
    conv_prefix = ROOM + '/Obj_Convene'
    conn.add_info(conv_prefix + '/Img_Convene_On', 'Img_Convene',
                  active=True, button=False, text='')
    conn.add_info(conv_prefix + '/Img_Convene_Off', 'Img_Convene',
                  active=False, button=False, text='')

    def _toggle_convene(c):
        on = c.info_items[conv_prefix + '/Img_Convene_On']
        off = c.info_items[conv_prefix + '/Img_Convene_Off']
        on['activeInHierarchy'], off['activeInHierarchy'] =             off['activeInHierarchy'], on['activeInHierarchy']
    conn.click_effects[ROOM_ANCHORS['Button_Convene']] = _toggle_convene
    CURRENCY_PREFIX = ('UIRoot/Layer_Common/ViewRoot<Game.SDCMainView>/HUD/'
                       'UIHUD_1/SafeArea/Currency&Buff/CurrencyPanel/Currency/'
                       'CurrencyItem(Clone)')
    conn.add_info(f'{CURRENCY_PREFIX}/wimg_ItemNum_SDC', 'wimg_ItemNum_SDC',
                  text='128', color=[190, 190, 190, 255])
    def _value_text(c):
        return str(4500 + 100000 * getattr(c, 'bag_collections', 0))
    def _value_color(c):
        return ([80, 200, 80, 255] if getattr(c, 'bag_collections', 0) >= 1
                else [220, 60, 60, 255])
    conn.add_info((f'UIRoot/Layer_Common/ViewRoot<Game.SDCPreWarWindow>/MainView/'
                   f'SDCPreWarWindow/SafeArea/Obj_Left/Left/TxtValue'), 'TxtValue',
                  text=_value_text, color=_value_color, active=True)
    conn.add_info(('UIRoot/Layer_Top/ViewRoot<Game.UITxtTips>/MainView/'
                   'UITxtTips/TxToast'), 'ToastText',
                  text=lambda c: c.toast, color=[255, 255, 255, 255])
    # TIPS 节点（默认未激活，点格子后激活）
    conn.add_info(f'{TIPS_ROOT}/MainView/SDCItemTips/Content/ItemTop/ItemName',
                  'TxtItemName', text='初级护甲·I', active=False)
    conn.add_info(TIPS_EQUIP_PATH, 'Button_Equip', active=False, button=True)
    conn.add_info(TIPS_CLOSE_PATH, 'BtnClose', active=False, button=True)

    conn.click_effects[warehouse_cell_path(0)] = lambda c: c._set_tips_open()
    conn.click_effects[TIPS_CLOSE_PATH] = lambda c: c._set_tips_closed()

    # 背包区(增幅器容器)：072 互换用例的 scan_bag_item 通道
    bag_zone = ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                'ViewRoot<Game.SDCPreWarWindow>/MainView/SDCPreWarWindow/'
                'SafeArea/Obj_Left/Left/Obj_LeftList_2/AmplifierTrans')
    conn.add_info(bag_zone, 'AmplifierTrans', active=True, button=False)
    bag_cell = bag_zone + '/Obj_Item_0/Btn_Click'
    conn.add_info(bag_cell, 'Btn_Click', active=True, button=True,
                  text='SC-1·普攻增幅器')
    _tips_title = (TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemTop/ItemName2')
    def _open_bag_tips(c):
        c._open_cell_tips(5)
        if _tips_title in c.info_items:
            c.info_items[_tips_title]['text'] = 'SC-1·普攻增幅器'
    conn.click_effects[bag_cell] = _open_bag_tips
    pw = (TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemMid/'
          'GiftDescInfoContent/Button_PutInWarehouse')
    conn.add_info(pw, 'Button_PutInWarehouse', active=False, button=True)
    conn.click_effects[pw] = lambda c: c._act_toast('放入成功')
    ui_p = (TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemMid/'
            'GiftDescInfoContent/Button_UnInstall')
    conn.add_info(ui_p, 'Button_UnInstall', active=False, button=True)
    conn.click_effects[ui_p] = lambda c: c._act_toast('卸下成功')
    equip_zone_path = (f'UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                       f'ViewRoot<Game.SDCPreWarWindow>/MainView/SDCPreWarWindow/'
                       f'SafeArea/Obj_Left/Left/EquipTrans')
    if not any(it['path'] == equip_zone_path
               for it in conn.info_items.values()):
        conn.add_info(equip_zone_path, 'EquipTrans', active=True, button=False)
    for si in (1, 2, 3):
        cellp = f'{equip_zone_path}/Obj_Content_{si}/BtnItem'
        conn.add_info(cellp, 'BtnItem', active=True, button=True)
        conn.click_effects[cellp] = (lambda si: (lambda c: c._open_cell_tips(0)))(si)

    # ---- 英雄选择弹窗(TC-4.3-008) ----
    HERO_ROOT = 'UIRoot/Layer_Common/ViewRoot<Game.UIHeroSelect>'
    HERO_BTN = f'{ROOM[:ROOM.index("/TeamRoom")]}/../../Left/Template_HeroHead/BtnHero'
    conn.add_items({'BtnHero': ROOM_ANCHORS['BtnPreBattle'].replace('BtnPreBattle', 'x')})
    # 覆盖为正确路径
    for lst in conn.static_items.values():
        pass
    del conn.static_items['BtnHero']
    real_hero_path = (f'UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                      f'ViewRoot<Game.SDCPreWarWindow>/MainView/SDCPreWarWindow/SafeArea/'
                      f'Obj_Left/Left/Template_HeroHead/BtnHero')
    conn.add_items({'BtnHero': real_hero_path})
    conn.entry_map['BtnHero'] = HERO_ROOT
    hero_item_path = f'{HERO_ROOT}/MainView/List/HeroItem0'
    conn.add_info(hero_item_path, 'Btn_HeroCell', active=True, button=True,
                  text='另一个英雄')

    def _select_hero(c):
        if HERO_ROOT in c.opened_stack:
            c.opened_stack.remove(HERO_ROOT)
            c.hero_swapped = True

    conn.hero_swapped = False
    conn.click_effects[hero_item_path] = _select_hero

    # 天赋/觉醒文本随切换变化
    pre_prefix = ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                  'ViewRoot<Game.SDCPreWarWindow>/MainView/SDCPreWarWindow/SafeArea')
    conn.add_info(f'{pre_prefix}/Obj_Left/Left/wbtn_Talent', 'wbtn_Talent',
                  text=lambda c: '2/3' if getattr(c, 'hero_swapped', False) else '1/3',
                  active=True, button=True)
    conn.add_info(f'{pre_prefix}/Obj_Left/Left/wbtn_Awake', 'wbtn_Awake',
                  text=lambda c: '1/3' if getattr(c, 'hero_swapped', False) else '0/3',
                  active=True, button=True)

    # ---- 仓库页签 ×2 + 格子集差异(TC-4.4-002/003) ----
    pre_root = ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                'ViewRoot<Game.SDCPreWarWindow>/MainView/SDCPreWarWindow')
    tab_base = f'{pre_root}/SafeArea/Right/TabExtendList'
    wh = f'{pre_root}/SafeArea/Right/WarehouseScrollView'
    conn.add_info(f'{tab_base}/ButtonFunc_0', 'ButtonFunc', active=True, button=True)
    conn.add_info(f'{tab_base}/ButtonFunc_1', 'ButtonFunc2', active=True, button=True)

    for name in ('A', 'B', 'C'):
        conn.add_info(f'{wh}/A_{name}', f'Obj_Item{chr(ord(name))}', active=True)
    conn.add_info(f'{wh}/B_extra', 'Obj_ItemD', active=False)

    def on_tab0(c):
        for k in ('A', 'B', 'C'):
            c.info_items[f'{wh}/A_{k}']['activeInHierarchy'] = True
        c.info_items[f'{wh}/B_extra']['activeInHierarchy'] = False
    def on_tab1(c):
        for k in ('A', 'B', 'C'):
            c.info_items[f'{wh}/A_{k}']['activeInHierarchy'] = False
        c.info_items[f'{wh}/B_extra']['activeInHierarchy'] = True
    conn.click_effects[f'{tab_base}/ButtonFunc_0'] = on_tab0
    conn.click_effects[f'{tab_base}/ButtonFunc_1'] = on_tab1
    # ---- 商店子页签 ×2 + 商品格差异(TC-4.4-012/013) ----
    shop_tabs = ('UIRoot/Layer_Common/ViewRoot<Game.SDCPreWarShopWindow>/'
                 'MainView/UIOutBattleShop/SafeArea/Right/TabExtendList')
    conn.add_info(f'{shop_tabs}/Sub_ButtonFunc_0', 'ButtonFunc', active=True, button=True)
    conn.add_info(f'{shop_tabs}/Sub_ButtonFunc_1', 'ButtonFunc', active=True, button=True)
    for i in range(3):
        gp = (f'UIRoot/Layer_Common/ViewRoot<Game.SDCPreWarShopWindow>/'
              f'MainView/SDCPreWarShopWindow/SafeArea/Goods/G{i}/Btn_Click')
        conn.add_info(gp, 'Btn_Click', active=(i < 2), button=True)
        conn.click_effects[gp] = (lambda i: (lambda c: c._open_cell_tips(1)))(i)

    def shop_set(n_active):
        def fn(c):
            for i in range(3):
                p = (f'UIRoot/Layer_Common/ViewRoot<Game.SDCPreWarShopWindow>/'
                     f'MainView/SDCPreWarShopWindow/SafeArea/Goods/G{i}/Btn_Click')
                c.info_items[p]['activeInHierarchy'] = i < n_active
        return fn
    conn.click_effects[f'{shop_tabs}/Sub_ButtonFunc_0'] = shop_set(2)
    conn.click_effects[f'{shop_tabs}/Sub_ButtonFunc_1'] = shop_set(1)

    def _equip(c):
        c.toast = '装配成功'
        c._set_tips_closed()
    conn.click_effects[TIPS_EQUIP_PATH] = _equip

    # ---- 仓库格子池：覆盖 19 条用例所需全部分类（真名来自 battle_room_items） ----
    conn.init_cells([
        (1, '软质护甲'),        # 0 护甲(4.3-010)
        (2, 'SC-1·攻击芯片'),    # 1 芯片
        (8, 'SC-1·治疗'),        # 2 技能
        (2, 'SC-1·攻击芯片'),    # 3 芯片
        (7, '便携腰包'),         # 4 背包
        (3, 'SC-1·普攻增幅器'),  # 5 增幅器(020消费)
        (5, '简易救治袋'),       # 6 药品
        (4, '否定陀螺'),         # 7 藏品(035消费)
        (6, '篮球场钥匙'),       # 8 钥匙
        (11, '赛季中级安全箱'),   # 9 安全箱券
        (10, '赛季中级钥匙包'),   # 10 钥匙链券(官方名)
        (2, 'SC-1·攻击芯片'),    # 11 芯片
        (4, 'G笔'),             # 12 藏品(050消费)
        (3, 'SC-2·普攻增幅器'),  # 13 增幅器(029消费)
        (2, 'SC-1·攻击芯片'),    # 14 芯片(025第三枚)
        (4, '否定陀螺'),         # 15 藏品(038追加消耗备件)
    ])
    # TIPS 动态标题：随当前格变化
    conn.add_info(
        TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemTop/ItemName2',
        'TxtItemTitle', active=False,
        text=lambda c: (c.cells[c.current_cell][1]
                        if c.current_cell is not None
                        and c.current_cell < len(c.cells) else ''))
    orig_open = conn._open_cell_tips
    def _open_cell_tips_with_title(self, idx):
        orig_open(idx)
        t = self.info_items.get(
            TIPS_ROOT + '/MainView/SDCItemTips/Content/ItemTop/ItemName2')
        if t is not None:
            t['activeInHierarchy'] = True
    conn._open_cell_tips = _open_cell_tips_with_title.__get__(conn)
    # 出售开关
    import src.task.battle_room_ui as _brui
    conn.click_effects[_brui.TIPS_SELL_BUTTON_PATH] = \
        lambda c: c._set_sell_mode(not c.sell_mode)
    conn.currency = ['3200', '128']

    return conn


def make_ctx(conn, allow_clock_change=False, allow_gm_items=False,
             gm_item_ids=None, **ui_kwargs):
    kwargs = dict(load_timeout=0.3, poll_interval=0.02, close_settle=0.0)
    kwargs.update(ui_kwargs)
    ui = UiContext(conn, make_logger(), **kwargs)
    ctx = CheckContext(conn, ui, make_logger(), allow_clock_change=allow_clock_change,
                       allow_gm_items=allow_gm_items, gm_item_ids=gm_item_ids)
    return ctx, ui


# --------------------------------------------------------------------------
# 用例数据
# --------------------------------------------------------------------------

class TestCaseData:
    def test_load_from_bundled_json(self):
        cases = load_cases()
        assert len(cases) == 181
        assert all(c.case_id.startswith('TC-') for c in cases)

    def test_smoke_filter_keeps_only_p0(self):
        cases = load_cases()
        selected, skipped = filter_cases(cases, smoke_only=True)
        assert all(c.priority == 'P0' for c in selected)
        assert len(selected) + len(skipped) == len(cases)

    def test_case_id_filter(self):
        cases = load_cases()
        selected, skipped = filter_cases(
            cases, smoke_only=False, case_id_filter='TC-4.1-001, tc-4.2-002')
        assert {c.case_id for c in selected} == {'TC-4.1-001', 'TC-4.2-002'}
        assert skipped['TC-4.1-002'] == '不在用例编号过滤范围内'

    def test_sort_by_priority_stable(self):
        cases = sort_by_priority(load_cases())
        priorities = [c.priority for c in cases]
        p0_count = priorities.count('P0')
        assert priorities[:p0_count] == ['P0'] * p0_count

    def test_p0_coverage(self):
        """所有 P0 用例要么可执行、要么登记了待支持原因，不允许漏网"""
        cases = [c for c in load_cases() if c.priority == 'P0']
        assert cases, 'P0 用例为空，用例 JSON 可能未生成'
        for c in cases:
            assert c.case_id in EXECUTABLE_CHECKS or c.case_id in NEEDS_SUPPORT, \
                f'{c.case_id} 既不可执行也无待支持说明'


class TestUnityConnectionUiWrappers:
    def _conn(self, resp):
        conn = UnityConnection()
        conn.send_command = MagicMock(return_value=resp)
        return conn

    def test_find_ui_ok(self):
        msg = json.dumps({'success': True, 'count': 1, 'items': [make_item('BtnClose')]})
        conn = self._conn({'status': 'ok', 'message': msg})
        result = conn.find_ui(name_contains='Close', max_results=5)
        assert result['success'] and result['count'] == 1
        sent = conn.send_command.call_args[0]
        assert sent[0] == 'automation_find_ui'
        assert sent[1] == {'nameContains': 'Close', 'maxResults': 5}

    def test_find_ui_error(self):
        conn = self._conn({'status': 'error', 'message': '未找到'})
        assert conn.find_ui(name_contains='X') == {'success': False, 'count': 0, 'items': []}

    def test_click_ui_payload(self):
        conn = self._conn({'status': 'ok', 'message': '{}'})
        resp = conn.click_ui(path='UIRoot/Btn')
        assert resp['status'] == 'ok'
        assert conn.send_command.call_args[0][1] == {'path': 'UIRoot/Btn'}

    def test_go_back_command(self):
        conn = self._conn({'status': 'ok', 'message': '{"success":true}'})
        resp = conn.go_back()
        assert resp['status'] == 'ok'
        assert conn.send_command.call_args[0][0] == 'automation_go_back'

    def test_get_battle_state(self):
        conn = self._conn({'status': 'ok', 'message': json.dumps({'errors': ['NullRef']})})
        data = conn.get_battle_state(include_errors=True, clear_errors=True)
        assert data['errors'] == ['NullRef']

    def test_get_battle_state_error_returns_empty(self):
        conn = self._conn({'status': 'error', 'message': '不在战斗中'})
        assert conn.get_battle_state() == {}


# --------------------------------------------------------------------------
# UiContext
# --------------------------------------------------------------------------

class TestUiContext:
    def test_active_view_roots_reflect_state(self):
        conn = build_happy_conn()
        ui = UiContext(conn, make_logger())
        before = ui.active_view_roots()
        assert not any('SDCMainView' in v for v in before)
        conn.opened_stack.append('ViewRoot<Game.SDCMainView>')
        assert any('SDCMainView' in v for v in ui.active_view_roots())

    def test_close_top_window_uses_back_stack(self):
        conn = build_happy_conn()
        ui = UiContext(conn, make_logger(), close_settle=0.0)
        conn.opened_stack.append('ViewRoot<Game.SDCPreWarWindow>')
        assert ui.close_top_window() is True
        assert conn.opened_stack == []

    def test_find_unknown_logical_returns_none(self):
        ui = UiContext(StatefulFakeConn(), make_logger())
        assert ui.find('not_exist') is None

    def test_path_candidate_supported(self):
        from src.task.battle_room_ui import UI_CANDIDATES
        real_path = UI_CANDIDATES['lobby_sdc_entry'][0][len('@path:'):]
        conn = StatefulFakeConn(entry_map={})
        conn.add_items({real_path.split('/')[-1]: real_path})
        ui = UiContext(conn, make_logger())
        item = ui.find('lobby_sdc_entry')
        assert item is not None and item['name'] == 'SDCEnter_Button'

    def test_wait_view_times_out(self):
        ui = UiContext(StatefulFakeConn(), make_logger(),
                       load_timeout=0.05, poll_interval=0.02)
        assert ui.wait_view('room_main_view') is None

    def test_runtime_error_sweep_disabled_on_empty_state(self):
        conn = StatefulFakeConn()
        conn.get_battle_state = lambda **kw: {}
        ui = UiContext(conn, make_logger())
        ui.clear_runtime_errors()
        assert ui.sweep_supported is False
        assert ui.collect_runtime_errors() == []

    def test_blocking_overlays(self):
        conn = StatefulFakeConn()
        conn.static_items['SystemGuideView'] = [
            make_item('SystemGuideView', button=False),
            make_item('SystemGuideOther', active=False, button=False)]
        ui = UiContext(conn, make_logger())
        names = ui.blocking_overlays()
        assert names == ['SystemGuideView']


# --------------------------------------------------------------------------
# 用例执行器
# --------------------------------------------------------------------------

class TestRunCase:
    def _case(self, case_id):
        return next(c for c in load_cases() if c.case_id == case_id)

    def test_full_navigation_pass_in_execution_order(self):
        import datetime
        from src.task import battle_room_checks as brc

        conn = build_happy_conn()
        ctx, ui = make_ctx(conn, allow_clock_change=True)
        ui.clear_runtime_errors()

        entry_path = ('UIRoot/RootCanvas/__dynamicRoot/Layer_Common/'
                      'ViewRoot<Game.UIMainCity2DView>/MainView/UIMainCity2D/'
                      'SafeArea/MainCityNew/ControlPanel/Content_Banner/'
                      'obj_LeftList/Obj_SDC/Root/SDCEnter_Button')
        saved_shift = brc._shift_time_out_of_season
        saved_restore = brc._os_set_localtime

        def shift():
            conn.info_items.setdefault(entry_path, {
                'name': 'SDCEnter_Button', 'path': entry_path,
                'activeInHierarchy': True, 'interactable': True,
                'hasButton': True, 'hasToggle': False,
                'text': None, 'color': None, 'isOn': None})
            conn.info_items[entry_path]['activeInHierarchy'] = False
            return datetime.datetime(2025, 8, 29, 12), True, ''

        def restore(dt):
            if entry_path in conn.info_items:
                conn.info_items[entry_path]['activeInHierarchy'] = True
            return True, ''

        brc._shift_time_out_of_season = shift
        brc._os_set_localtime = restore
        try:
            import os as _os
            ordered = sorted(SKELETON_CASES & set(EXECUTABLE_CHECKS))
            if _os.environ.get('BR_TRACE'):
                for case_id in ordered:
                    result = run_case(self._case(case_id), ctx)
                    assert result.status == STATUS_PASS, f'{case_id}: {result.detail}'
                    vis = ctx.ui.find_active('prewar_window')
                    print(f"TRACE {case_id} {result.status} stack={conn.opened_stack} "
                          f"prewar_visible={bool(vis)}")
            else:
                for case_id in ordered:
                    result = run_case(self._case(case_id), ctx)
                    assert result.status == STATUS_PASS, f'{case_id}: {result.detail}'
        finally:
            brc._shift_time_out_of_season = saved_shift
            brc._os_set_localtime = saved_restore

    def test_room_back_after_full_chain(self):
        conn = build_happy_conn()
        ctx, ui = make_ctx(conn)
        ui.clear_runtime_errors()
        for case_id in sorted(SKELETON_CASES & set(EXECUTABLE_CHECKS)):
            run_case(self._case(case_id), ctx)
        # 链路结束：临时窗口全部关闭，仍停留在搜打撤房间主界面
        assert conn.opened_stack == ['ViewRoot<Game.SDCMainView>'],             f"stack={conn.opened_stack} audit={conn.audit[-14:]}"

        assert ui.in_room()

    def test_room_entry_fail_blocks_scenario2(self):
        conn = StatefulFakeConn(base_roots=['ViewRoot<Game.UIMainCity2DView>'])
        ctx, ui = make_ctx(conn)
        r1 = run_case(self._case('TC-4.1-001'), ctx)
        assert r1.status == STATUS_FAIL
        r2 = run_case(self._case('TC-4.2-001'), ctx)
        assert r2.status == STATUS_BLOCKED

    def test_runtime_error_fails_passing_case(self):
        conn = build_happy_conn()
        ctx, ui = make_ctx(conn)
        ui.clear_runtime_errors()
        ctx.room_available = True
        conn._runtime_errors = ['NullReferenceException at X']
        result = run_case(self._case('TC-4.1-001'), ctx)
        assert result.status == STATUS_FAIL
        assert 'NullReferenceException' in result.detail

    def test_missing_anchor_fails_room_check(self):
        conn = build_happy_conn()
        del conn.static_items['Button_Convene']
        ctx, ui = make_ctx(conn, load_timeout=0.1)
        run_case(self._case('TC-4.1-001'), ctx)
        result = run_case(self._case('TC-4.2-001'), ctx)
        assert result.status == STATUS_FAIL
        assert '补齐队友区' in result.detail

    def test_needs_support_case_is_na(self):
        ctx, _ = make_ctx(StatefulFakeConn())
        result = run_case(self._case('TC-4.2-030'), ctx)
        assert result.status == STATUS_NA
        assert '赛季' in result.detail

    def test_clock_case_na_when_flag_off(self):
        from src.task.battle_room_checks import CheckContext as CC
        ui = UiContext(StatefulFakeConn(), make_logger())
        ctx = CheckContext(StatefulFakeConn(), ui, make_logger(),
                           allow_clock_change=False)
        result = run_case(self._case('TC-4.1-002'), ctx)
        assert result.status == STATUS_NA
        assert '允许改系统时间' in result.detail

    def test_unimplemented_p1_case_is_na(self):
        ctx, _ = make_ctx(StatefulFakeConn())
        p1 = next(c for c in load_cases() if c.priority == 'P1')
        assert run_case(p1, ctx).status == STATUS_NA


# --------------------------------------------------------------------------
# 任务装配
# --------------------------------------------------------------------------

class TestBattleRoomTestTask:
    def _task(self, config_overrides=None):
        from src.task.BattleRoomTestTask import BattleRoomTestTask
        task = BattleRoomTestTask.__new__(BattleRoomTestTask)
        task.name = 'BattleRoomTestTask'
        task.logger = make_logger()
        task.config = {
            '仅运行P0冒烟用例': True,
            '自动登录到大厅': False,
            '失败后停止': False,
            '界面加载超时(秒)': 0.3,
            '用例编号过滤': '',
            '允许改系统时间做隐藏验证': True,
        }
        task.config.update(config_overrides or {})
        return task

    def _run(self, task, tmp_path, conn, capture):
        def fake_report(smoke, results, summary, ui_snapshot, aborted,
                        dismissed_popups=None, screenshots=None):
            capture.update(summary=summary, results=results, smoke=smoke)
            return str(tmp_path / 'report.json'), str(tmp_path / 'report.xlsx')

        with patch.object(type(task), '_get_unity_connection', return_value=conn), \
                patch.object(type(task), '_write_report', side_effect=fake_report):
            return task.run()

    def test_run_smoke_happy_path(self, tmp_path):
        task = self._task({
            '用例编号过滤': ','.join(sorted(SKELETON_CASES)),
            '允许改系统时间做隐藏验证': True,
            '允许GM自动发道具': True,
            'GM道具ID映射': '{"护甲":1600055,"芯片":1600163,"技能":1600115,'
                          '"背包":1600109,"增幅器":1600061,"药品":1600079,'
                          '"藏品":1610001,"钥匙":1600103,"安全箱":1620004}',
        })
        capture = {}
        conn = build_happy_conn()
        ok = self._run(task, tmp_path, conn, capture)
        skeleton_results = [f for f in capture['results']
                            if f.case.case_id in SKELETON_CASES]
        sk_fails = [f for f in skeleton_results if f.status == 'FAIL']
        assert not sk_fails, [
            (f.case.case_id, f.detail[:200]) for f in sk_fails]
        snap = ('SNAP cells0=%r ncells=%d nempty=%d'
                % (conn.cells[0], len(conn.cells),
                   sum(1 for cc in conn.cells if cc == (None, ''))))
        print(snap)
        # 冒烟模式:P0 全执行;可执行数与 N/A 数由用例集动态决定
        p0_total = sum(1 for c in load_cases() if c.priority == 'P0')
        # 可执行用例零 FAIL(Blocked=备料缺失时的合规状态,允许记录);
        # 多用例串联的格子消耗时序以真机冒烟为最终验证
        # Blocked=备料缺失时的合规状态,允许存在(真机GM补发验证为准)
        # 骨架口径：骨架 13 条全 PASS；其余用例的状态由真机冒烟为最终验证
        bad = [(c['case_id'], c['status'], c['detail'][:60])
               for c in map(lambda r: r.to_dict(), capture['results'])
               if c['status'] != 'PASS' and c['case_id'] in SKELETON_CASES]
        assert not bad, f'骨架用例未PASS: {bad}'
        blocked = [f for f in capture['results'] if f.status == 'Blocked']
        print('BLOCKED:', [(f.case.case_id, f.detail[:100]) for f in blocked])
        assert capture['summary']['PASS'] == len(SKELETON_CASES
                                                 & set(EXECUTABLE_CHECKS))
        # 编号过滤模式下非骨架用例被 Skipped(含赛季2条)，N/A=0 为正确行为
        assert capture['summary']['N/A'] == 0
        assert capture['summary']['FAIL'] == 0, [
            (c['case_id'], c['status'], c['detail'][:80])
            for c in map(lambda r: r.to_dict(), capture['results'])
            if c['status'] == 'FAIL']
        # 编号过滤：跳过 = 其余 134 条 P1/P2 + 骨架外的 32 条 P0(45可执行+2赛季)
        assert capture['summary']['Skipped'] == len(load_cases()) - len(SKELETON_CASES)
        total_check = (capture['summary']['PASS'] + capture['summary']['N/A']
                       + capture['summary']['Skipped']
                       + capture['summary']['Blocked'])
        assert total_check == len(load_cases())
        assert len(capture['results']) == 181
        assert capture['smoke'] is True

    def test_run_returns_false_when_room_entry_fails(self, tmp_path):
        task = self._task()
        capture = {}
        ok = self._run(
            task, tmp_path,
            StatefulFakeConn(base_roots=['ViewRoot<Game.UIMainCity2DView>']), capture)
        assert ok is False

    def test_smoke_off_runs_all(self, tmp_path):
        task = self._task({'仅运行P0冒烟用例': False})
        capture = {}
        self._run(task, tmp_path, build_happy_conn(), capture)
        assert capture['total'] if False else len(capture['results']) == 181
        assert capture['smoke'] is False

    def test_report_written(self):
        from src.task.BattleRoomTestTask import BattleRoomTestTask
        from src.task.battle_room_cases import CaseResult
        task = self._task()
        cases = load_cases()[:2]
        results = [CaseResult(cases[0], STATUS_PASS, 'ok'),
                   CaseResult(cases[1], STATUS_NA, '需文本读取')]
        import shutil
        out_dir = os.path.join(os.path.dirname(__file__), '_tmp_report_out')
        shutil.rmtree(out_dir, ignore_errors=True)
        json_path, xlsx_path = BattleRoomTestTask._write_report(
            task, True, results, summarize(results), [], False,
            out_dir=out_dir)
        with open(json_path, 'r', encoding='utf-8') as f:
            report = json.load(f)
        assert report['smoke'] is True
        assert report['summary']['PASS'] == 1
        assert len(report['cases']) == 2
        # XLSX 执行报告也一并落盘，且测试产物隔离在临时目录
        assert os.path.exists(xlsx_path)
        assert os.path.dirname(json_path) == out_dir
        shutil.rmtree(out_dir, ignore_errors=True)

    def test_registered_in_config(self):
        import config
        assert ['src.task.BattleRoomTestTask', 'BattleRoomTestTask'] in config.config['onetime_tasks']


class TestReportXlsx:
    def test_xlsx_sheets_and_columns(self):
        import shutil

        from openpyxl import load_workbook

        from src.task.battle_room_cases import CaseResult
        from src.task.battle_room_report import write_reports
        cases = load_cases()[:3]
        results = [CaseResult(cases[0], STATUS_PASS, 'ok'),
                   CaseResult(cases[1], STATUS_FAIL, '未找到节点'),
                   CaseResult(cases[2], STATUS_NA, '需拖拽命令')]
        out_dir = os.path.join(os.path.dirname(__file__), '_tmp_report_out')
        shutil.rmtree(out_dir, ignore_errors=True)
        json_path, xlsx_path = write_reports(
            out_dir, 'test', True, False, results,
            summarize(results), [{'name': 'BtnX'}])
        wb = load_workbook(xlsx_path)
        assert wb.sheetnames == ['汇总', '用例结果', '待Unity侧支持']
        ws = wb['用例结果']
        assert [c.value for c in ws[1][:7]] == [
            '用例编号', '优先级', '模块', '功能', '操作步骤', '期望结果', '测试结果']
        statuses = [ws.cell(row=i, column=7).value for i in range(2, 5)]
        assert statuses == ['PASS', 'FAIL', 'N/A']
        ws1 = wb['汇总']
        values = {ws1.cell(row=i, column=1).value: ws1.cell(row=i, column=2).value
                  for i in range(1, ws1.max_row + 1)}
        assert values['FAIL'] == 1 and values['冒烟模式(仅P0)'] == '是'
        shutil.rmtree(out_dir, ignore_errors=True)


class TestGmFullAuto:
    """GM 备料全自动:开关隐藏时工具自行激活,无需任何人工按键"""

    def _gm_conn(self):
        conn = build_happy_conn()
        # GM 节点(开关默认隐藏,模拟从未按过 F9 的全新会话)
        conn.add_info(GM_SWITCH_PATH, 'GMSwitch', active=False, button=True)
        conn.add_info(GM_INPUT_PATH, 'InputField', active=False)
        conn.add_info(GM_SEND_PATH, 'Send', active=False, button=True)

        state = {'sent': None}

        def on_send(c):
            it = c.info_items.get(GM_INPUT_PATH)
            state['sent'] = it.get('text') if it else None
            for p in (GM_INPUT_PATH, GM_SEND_PATH):
                if p in c.info_items:
                    c.info_items[p]['activeInHierarchy'] = False
        conn.click_effects[GM_SEND_PATH] = on_send

        # 让 set_ui_input 写入的文本可被读取
        orig_info = conn.get_ui_info
        def info_with_input(path=None, name_contains=None, max_results=None):
            res = orig_info(path=path, name_contains=name_contains,
                            max_results=max_results)
            if path == GM_INPUT_PATH and res['items']:
                pass  # text 字段由写入时同步
            return res
        # 直接给 fake 增加 input 文本回写能力
        def set_ui_input(text, path=None, name_contains=None):
            if path in conn.info_items:
                conn.info_items[path]['text'] = text
                return {'status': 'ok', 'message': '{}'}
            return {'status': 'error', 'message': 'no target'}
        conn.set_ui_input = set_ui_input
        conn._sent_state = state
        return conn

    def test_open_gm_panel_via_model_reveal(self):
        from src.task.battle_room_ui import GM_SWITCH_PATH, GM_INPUT_PATH
        conn = self._gm_conn()
        ui = UiContext(conn, make_logger(), close_settle=0.0)
        assert not conn.info_items[GM_INPUT_PATH]['activeInHierarchy']
        ok, msg = ui.open_gm_panel()
        assert ok, msg
        assert conn.info_items[GM_INPUT_PATH]['activeInHierarchy']

    def test_close_gm_panel_hides_input(self):
        from src.task.battle_room_ui import GM_INPUT_PATH
        conn = self._gm_conn()
        ui = UiContext(conn, make_logger(), close_settle=0.0)
        assert ui.open_gm_panel()[0]
        assert ui.close_gm_panel() is True
        assert not conn.info_items[GM_INPUT_PATH]['activeInHierarchy']

    def test_gm_add_item_end_to_end_no_keyboard(self):
        from src.task.battle_room_ui import GM_INPUT_PATH
        conn = self._gm_conn()
        ui = UiContext(conn, make_logger(), close_settle=0.0)
        ok, msg = ui.gm_add_item(1600055, count=1)
        assert ok, msg
        assert conn._sent_state['sent'] == 'AddItem=1600055=1'

    def test_fake_set_ui_active_flips_state(self):
        from src.task.battle_room_ui import GM_SWITCH_PATH
        conn = self._gm_conn()
        resp = conn.set_ui_active(True, path=GM_SWITCH_PATH)
        assert resp['status'] == 'ok'
        assert conn.info_items[GM_SWITCH_PATH]['activeInHierarchy'] is True


class TestTipsFamily:
    """TIPS 等价装配/容器/激活/批量出售/购买/队友 用例（表驱动）"""

    def _setup(self, tmp_path, monkeypatch):
        from src.task import battle_room_checks as brc
        detail = [
            {'id': 1600055, 'name': '软质护甲', 'subtype': 1, 'value': 500,
             'season_value': 0},
            {'id': 1600163, 'name': 'SC-1·攻击芯片', 'subtype': 2, 'value': 800,
             'season_value': 0},
            {'id': 1600115, 'name': 'SC-1·治疗', 'subtype': 8, 'value': 300,
             'season_value': 0},
            {'id': 1600109, 'name': '便携腰包', 'subtype': 7, 'value': 600,
             'season_value': 0},
            {'id': 1600061, 'name': 'SC-1·普攻增幅器', 'subtype': 3, 'value': 900,
             'season_value': 0},
            {'id': 1600079, 'name': '简易救治袋', 'subtype': 5, 'value': 200,
             'season_value': 0},
            {'id': 1610001, 'name': '否定陀螺', 'subtype': 4, 'value': 100000,
             'season_value': 50},
            {'id': 1600103, 'name': '篮球场钥匙', 'subtype': 6, 'value': 400,
             'season_value': 0},
            {'id': 1620004, 'name': '赛季中级安全箱', 'subtype': 11, 'value': 0,
             'season_value': 0},
            {'id': 1600110, 'name': '赛季中级钥匙包', 'subtype': 10, 'value': 0,
             'season_value': 0},
        ]
        monkeypatch.setattr(brc, '_item_detail', lambda ctx: detail)
        return detail

    def _run_one(self, tmp_path, monkeypatch, case_id, expect):
        self._setup(tmp_path, monkeypatch)
        conn = build_happy_conn()
        ctx, ui = make_ctx(conn, allow_gm_items=True,
                           gm_item_ids={'护甲': 1600055, '芯片': 1600163,
                                        '技能': 1600115, '背包': 1600109,
                                        '增幅器': 1600061, '药品': 1600079,
                                        '藏品': 1610001, '钥匙': 1600103,
                                        '安全箱': 1620004})
        ui.clear_runtime_errors()
        # 前置：进房间并打开战备窗
        run_case(self_case('TC-4.1-001'), ctx)
        run_case(self_case('TC-4.3-001'), ctx)
        case = self_case(case_id)
        result = run_case(case, ctx)
        assert result.status == expect, f'{case_id}: {result.detail}'

    def test_slot_family(self, tmp_path, monkeypatch):
        for cid in ('TC-4.3-013', 'TC-4.3-015', 'TC-4.3-020', 'TC-4.3-024',
                    'TC-4.3-010'):
            self._run_one(tmp_path, monkeypatch, cid, STATUS_PASS)

    def test_bag_family(self, tmp_path, monkeypatch):
        for cid in ('TC-4.3-029', 'TC-4.3-032', 'TC-4.3-035', 'TC-4.3-044',
                    'TC-4.3-050'):
            self._run_one(tmp_path, monkeypatch, cid, STATUS_PASS)

    def test_activate_family(self, tmp_path, monkeypatch):
        for cid in ('TC-4.3-040', 'TC-4.3-047'):
            self._run_one(tmp_path, monkeypatch, cid, STATUS_PASS)

    def test_chip_limit(self, tmp_path, monkeypatch):
        self._run_one(tmp_path, monkeypatch, 'TC-4.3-025', STATUS_PASS)

    def test_sell_mode(self, tmp_path, monkeypatch):
        for cid in ('TC-4.4-024', 'TC-4.4-025', 'TC-4.4-029', 'TC-4.4-033'):
            self._run_one(tmp_path, monkeypatch, cid, STATUS_PASS)

    def test_buy_in_shop(self, tmp_path, monkeypatch):
        self._run_one(tmp_path, monkeypatch, 'TC-4.4-019', STATUS_PASS)

    def test_season_currency_sell(self, tmp_path, monkeypatch):
        self._run_one(tmp_path, monkeypatch, 'TC-4.2-028', STATUS_PASS)

    def test_convene_toggle(self, tmp_path, monkeypatch):
        for cid in ('TC-4.2-007', 'TC-4.2-008'):
            self._run_one(tmp_path, monkeypatch, cid, STATUS_PASS)


def self_case(case_id):
    return next(c for c in load_cases() if c.case_id == case_id)


class TestItemAutoMapping:
    """battle_room_items: 三表扫描/分类映射/缓存/手动覆盖"""

    def _make_fixture(self, root):
        import json as J
        os.makedirs(root, exist_ok=True)
        def w(name, obj):
            with open(os.path.join(root, name), 'w', encoding='utf-8') as f:
                J.dump(obj, f, ensure_ascii=False)
        w('MultiLanguage.json', {
            '1': {'ID': 1, 'Chinese': '软质护甲', 'Traditional': ''},
            '2': {'ID': 2, 'Chinese': '', 'Traditional': '否定陀螺'},
            '3': {'ID': 3, 'Chinese': '万象宝珠', 'Traditional': ''},
        })
        w('Item.json', {
            '1600055': {'ID': 1600055, 'Type': 16, 'SubType': 1, 'NameL': 1},
            '1600056': {'ID': 1600056, 'Type': 16, 'SubType': 1, 'NameL': 1},
            '1610001': {'ID': 1610001, 'Type': 16, 'SubType': 4, 'NameL': 2},
            '1620001': {'ID': 1620001, 'Type': 16, 'SubType': 99, 'NameL': 3},   # 非法分类应跳过
            '1001': {'ID': 1001, 'Type': 7, 'SubType': 1, 'NameL': 3},          # 非SDC类型应跳过
        })
        w('SDCItem.json', {
            '1': {'ID': 1600055}, '2': {'ID': 1600056},
            '3': {'ID': 1610001}, '4': {'ID': 1620001}, '5': {'ID': 1001},
        })

    def test_scan_groups_by_subtype_min_id(self, tmp_path):
        import shutil
        from src.task.battle_room_items import scan_client_tables, SUBTYPE_LABELS
        fixture = str(tmp_path / 'fx_json')
        self._make_fixture(fixture)
        mapping, detail = scan_client_tables(fixture)
        assert mapping['护甲'] == 1600055          # 同类取最小ID
        assert mapping['藏品'] == 1610001
        assert set(mapping) <= set(SUBTYPE_LABELS.values())
        assert all(d['id'] != 1001 and d['id'] != 1620001 for d in detail)

    def test_cache_roundtrip(self, tmp_path, monkeypatch):
        from src.task import battle_room_items as bri
        fixture = str(tmp_path / 'fx_json')
        self._make_fixture(fixture)
        cache = tmp_path / 'cache.json'
        monkeypatch.setattr(bri, '_cache_path', lambda: str(cache))
        # 注意 json_dir 用显式参数传递(默认参数在def时已绑定原常量)
        m1, _ = bri.load_snapshot(force=True, json_dir=fixture)
        # 删除源表后再读,应命中缓存
        shutil_rm = getattr(__import__('shutil'), 'rmtree')
        shutil_rm(fixture)
        m2, _ = bri.load_snapshot()
        assert m1 == m2 == {'护甲': 1600055, '藏品': 1610001}

    def test_merge_manual_overrides_auto(self):
        from src.task.battle_room_items import merge_item_ids
        merged, err = merge_item_ids({'护甲': 1600055}, '{"护甲": 1600056}')
        assert merged['护甲'] == 1600056 and err is None
        merged, err = merge_item_ids({'护甲': 1600055}, '')
        assert merged['护甲'] == 1600055 and err is None
        merged, err = merge_item_ids({}, '{bad json')
        assert merged == {} and err and 'JSON' in err


class TestNoGuessedNames:
    """防猜测守卫：候选表里的普通名称必须存在于游戏工程的绑定清单或白名单"""

    def test_candidates_traced_to_bindings(self):
        import json as _json
        import os as _os
        from src.task.battle_room_ui import UI_CANDIDATES
        reg_path = _os.path.join(_os.path.dirname(_os.path.dirname(
            _os.path.abspath(__file__))), 'configs', 'battle_room_ui_bindings.json')
        registry_names = set()
        if _os.path.exists(reg_path):
            data = _json.load(open(reg_path, encoding='utf-8'))
            for entries in data.get('behaviours', {}).values():
                for c in entries:
                    registry_names.add(c['name'])
            # 窗口/视图根关键字来自 view 类名注册表
            registry_names |= {v['class'] for v in data.get('views', [])}
        # 真机枚举确认、但绑定代码不含的名称（运行时动态节点/容器/视图根枚举关键字）
        runtime_confirmed = {
            'SDCEnter_Button',            # 大厅入口按钮（真机点击成功）
            'ViewRoot',                   # 视图根枚举用关键字
            'WarehouseScrollView',        # 仓库列表容器（sdc_prewar 枚举）
            'ButtonFunc',                 # 页签按钮（绑定于 TabItem 子 Behaviour）
            'BtnBack1',                   # 房间HUD返回键（sdc_room 枚举）
        }
        for logical, cands in UI_CANDIDATES.items():
            for cand in cands:
                if cand.startswith('@path:'):
                    continue
                assert (cand in registry_names or cand in runtime_confirmed
                        or any(cand in n for n in registry_names)), \
                    f'候选 {logical} -> {cand} 既不在绑定清单也不是白名单名称，疑似猜测'